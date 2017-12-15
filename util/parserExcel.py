#-*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import openpyxl
from openpyxl.styles import Border,Side,Font
import time


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red':'FFFF3030','green':'FF08B00'}


    def loadWorkBook(self,excelPathAndName):
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception,e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook


    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception,e:
            raise e


    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception,e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet


    def getRowsNumber(self,sheet):
        return sheet.max_row


    def getClosNumber(self,sheet):
        return sheet.max_column


    def getStartRowNumber(self,sheet):
        return sheet.min_row


    def getStartRowNubmer(self,sheet):
        return sheet.min_row


    def getRow(self,sheet,rowNo):
        try:
            return sheet.rows[rowNo-1]
        except Exception,e:
            raise e


    def getColumn(self,sheet,colNo):
        try:
            return sheet.col_values(colNo-1)
        except Exception,e:
            raise e


    def getCellOfValue(self,sheet,rowNo=None,colNo=None,coordinate=None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colNo).value
            except Exception,e:
                raise e
        else:
            raise Exception("error")


    def getCellOfObject(self,sheet,rowNo=None,colNo=None,coordinate=None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colNo)
            except Exception,e:
                raise e
        else:
            raise Exception("error")


    def writeCell(self,sheet,content,rowNo=None,colNo=None,coordinate=None,style=None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value =content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict(style))
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo,column=colNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo,column=colNo).font = Font(color=self.RGBDict(style))
                self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        else:
            raise Exception("error")


    def writeCellCurrentTime(self,sheet,rowNo=None,colNo=None,coordinate=None,style=None):
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value =currentTime

            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo,column=colNo).value = currentTime
            except Exception, e:
                raise e
        else:
            raise Exception("error")


if __name__ == "__main__":
    pe = ParseExcel()
    pe.loadWorkBook(u"C:\\Users\Administrator\\PycharmProjects\\TheLastKeyWord\\testData\\126邮箱发送邮件.xlsx")
    sheet = pe.getSheetByName("测试用例")
    # print pe.getRowsNumber(sheet)
    # print pe.getClosNumber(sheet)
    # rows = pe.getRow(sheet,1)
    # for temp in rows:
    #     print temp
    print sheet.rows